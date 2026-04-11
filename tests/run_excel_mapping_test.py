"""Experimental mapping test — runs the full pipeline for all 7 layouts.

Run from the repo root:
    python tests/run_excel_mapping_test.py

Requirements: openpyxl must be installed.
"""
from __future__ import annotations

import io
import sys
import traceback
from pathlib import Path

# Allow imports from src/
sys.path.insert(0, str(Path(__file__).parents[1] / "src"))

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter

from formbot.application.precision_fill import PrecisionFillUseCase
from formbot.domain.models import MappingRule


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def _make_wb_bytes(setup_fn) -> bytes:
    """Create an in-memory workbook, call setup_fn(wb) and return bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    setup_fn(wb, ws)
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _run_case(
    name: str,
    wb_bytes: bytes,
    rules: list[MappingRule],
    data: dict,
    expected: dict,  # {cell_ref: expected_value}
    min_confidence: float = 0.45,
) -> tuple[bool, str]:
    """
    Execute the full PrecisionFillUseCase pipeline.

    Returns (passed: bool, detail: str).
    """
    tmp = Path(sys.argv[0]).parent / "_tmp_excel_test"
    tmp.mkdir(exist_ok=True)
    template_path = tmp / f"{name}_template.xlsx"
    output_path = tmp / f"{name}_output.xlsx"

    template_path.write_bytes(wb_bytes)

    use_case = PrecisionFillUseCase(
        template_path=template_path,
        strict_mode=False,
        min_confidence=min_confidence,
        allow_overwrite_existing=False,
    )
    try:
        result = use_case.execute(data=data, mapping_rules=rules, output_path=output_path)
    except Exception as exc:
        return False, f"EXCEPTION during execute: {exc}\n{traceback.format_exc()}"
    finally:
        use_case.close()

    # Analyse decisions
    blocked = [d for d in result.decisions if d.status == "blocked"]
    if blocked:
        details = "; ".join(
            f"{d.field_name}@{d.label_position}→{d.target_position} reason={d.reason}"
            for d in blocked
        )
        return False, f"Blocked fields: {details}"

    # Verify cell values
    wb_out = openpyxl.load_workbook(output_path)
    try:
        ws = wb_out.active
        errors: list[str] = []
        for ref, expected_value in expected.items():
            actual = ws[ref].value
            if actual != expected_value:
                errors.append(f"  {ref}: expected={expected_value!r} actual={actual!r}")
        if errors:
            # Dump decisions for debugging
            decision_dump = "\n".join(
                f"    {d.field_name}: status={d.status} conf={d.confidence:.3f} "
                f"label_pos={d.label_position} target_pos={d.target_position} reason={d.reason}"
                for d in result.decisions
            )
            return False, "Cell value mismatches:\n" + "\n".join(errors) + "\n  Decisions:\n" + decision_dump
        return True, "OK"
    finally:
        wb_out.close()


# ---------------------------------------------------------------------------
# Layout definitions
# ---------------------------------------------------------------------------

def layout_a():
    """Layout A: label en col B, valor en col C (mismo row)."""
    def setup(wb, ws):
        ws["B2"] = "Razón social:"
        # C2 is empty — target

    rules = [
        MappingRule(
            field_name="razon_social",
            label="Razón social:",
            target_strategy="offset_or_infer",
            required=False,
        )
    ]
    data = {"razon_social": "EMPRESAS COLOMBIA SAS"}
    expected = {"C2": "EMPRESAS COLOMBIA SAS"}
    return _make_wb_bytes(setup), rules, data, expected


def layout_b():
    """Layout B: label en col A, valor en col B (mismo row)."""
    def setup(wb, ws):
        ws["A5"] = "NIT:"
        # B5 is empty — target

    rules = [
        MappingRule(
            field_name="nit",
            label="NIT:",
            target_strategy="offset_or_infer",
            required=False,
        )
    ]
    data = {"nit": "900123456-1"}
    expected = {"B5": "900123456-1"}
    return _make_wb_bytes(setup), rules, data, expected


def layout_c():
    """Layout C: label en col B row 5, valor en col B row 6 (abajo)."""
    def setup(wb, ws):
        ws["B5"] = "Dirección:"
        ws["C5"] = "ocupado"   # right side is occupied
        ws["D5"] = "ocupado"   # all rightward cells occupied up to scan limit
        ws["E5"] = "ocupado"
        ws["F5"] = "ocupado"
        ws["G5"] = "ocupado"
        ws["H5"] = "ocupado"
        ws["I5"] = "ocupado"
        ws["J5"] = "ocupado"
        # B6 is empty — target (below)

    rules = [
        MappingRule(
            field_name="direccion",
            label="Dirección:",
            target_strategy="offset_or_infer",
            required=False,
        )
    ]
    data = {"direccion": "Calle 10 # 5-30"}
    expected = {"B6": "Calle 10 # 5-30"}
    return _make_wb_bytes(setup), rules, data, expected


def layout_d():
    """Layout D: label con texto largo + celda vacía 2 cols a la derecha."""
    def setup(wb, ws):
        ws["A3"] = "Nombre completo del representante legal:"
        ws["B3"] = "ocupado"   # immediate right is occupied
        # C3 is empty — should be found (2 cols right)

    rules = [
        MappingRule(
            field_name="representante",
            label="Nombre completo del representante legal:",
            target_strategy="offset_or_infer",
            required=False,
        )
    ]
    data = {"representante": "JORGE ENRIQUE PÉREZ"}
    expected = {"C3": "JORGE ENRIQUE PÉREZ"}
    return _make_wb_bytes(setup), rules, data, expected


def layout_e():
    """Layout E: múltiples labels similares ('Teléfono 1', 'Teléfono 2')."""
    def setup(wb, ws):
        ws["A7"] = "Teléfono 1:"
        # B7 empty
        ws["A9"] = "Teléfono 2:"
        # B9 empty

    rules = [
        MappingRule(
            field_name="telefono_1",
            label="Teléfono 1:",
            target_strategy="offset_or_infer",
            required=False,
        ),
        MappingRule(
            field_name="telefono_2",
            label="Teléfono 2:",
            target_strategy="offset_or_infer",
            required=False,
        ),
    ]
    data = {"telefono_1": "601-3456789", "telefono_2": "601-9876543"}
    expected = {"B7": "601-3456789", "B9": "601-9876543"}
    return _make_wb_bytes(setup), rules, data, expected


def layout_f():
    """Layout F: merged cells en fila header, valor abajo en columna del label."""
    def setup(wb, ws):
        ws["B1"] = "Información del proveedor"
        ws.merge_cells("B1:E1")
        ws["B3"] = "Ciudad:"
        # C3 empty — target

    rules = [
        MappingRule(
            field_name="ciudad",
            label="Ciudad:",
            target_strategy="offset_or_infer",
            required=False,
        )
    ]
    data = {"ciudad": "Bogotá D.C."}
    expected = {"C3": "Bogotá D.C."}
    return _make_wb_bytes(setup), rules, data, expected


def layout_g():
    """Layout G: label 'NIT/CC:' — la '/' no debe confundir la búsqueda."""
    def setup(wb, ws):
        ws["A10"] = "NIT/CC:"
        # B10 empty — target

    rules = [
        MappingRule(
            field_name="nit_cc",
            label="NIT/CC:",
            target_strategy="offset_or_infer",
            required=False,
        )
    ]
    data = {"nit_cc": "800999888-7"}
    expected = {"B10": "800999888-7"}
    return _make_wb_bytes(setup), rules, data, expected


# ---------------------------------------------------------------------------
# P3 — offset=0,0 actually infers (doesn't write on label itself)
# ---------------------------------------------------------------------------

def p3_offset_zero_infers_correctly():
    """P3: row_offset=0, col_offset=0 → infiere hacia la derecha, no sobreescribe el label."""
    def setup(wb, ws):
        ws["A1"] = "Correo electrónico:"
        # B1 empty

    rules = [
        MappingRule(
            field_name="email",
            label="Correo electrónico:",
            row_offset=0,
            column_offset=0,
            target_strategy="offset_or_infer",
            required=False,
        )
    ]
    data = {"email": "test@empresa.com"}
    # Must NOT write to A1 (the label), must write to B1
    expected = {"A1": "Correo electrónico:", "B1": "test@empresa.com"}
    return _make_wb_bytes(setup), rules, data, expected


# ---------------------------------------------------------------------------
# P4 — celdas ya ocupadas (allow_overwrite_existing=False)
# ---------------------------------------------------------------------------

def p4_no_overwrite_occupied():
    """P4: celda adyacente ya tiene contenido → salta a la siguiente vacía."""
    def setup(wb, ws):
        ws["A1"] = "Banco:"
        ws["B1"] = "BANCO EXISTENTE"   # occupied
        # C1 is empty — should be used

    rules = [
        MappingRule(
            field_name="banco",
            label="Banco:",
            row_offset=0,
            column_offset=0,
            target_strategy="offset_or_infer",
            required=False,
        )
    ]
    data = {"banco": "BANCOLOMBIA"}
    # B1 must NOT be overwritten; C1 must receive the value
    expected = {"B1": "BANCO EXISTENTE", "C1": "BANCOLOMBIA"}
    return _make_wb_bytes(setup), rules, data, expected


# ---------------------------------------------------------------------------
# P5 — Labels con caracteres especiales
# ---------------------------------------------------------------------------

def p5_special_labels():
    """P5: labels con caracteres especiales — normalize_text debe encontrarlos."""
    def setup(wb, ws):
        ws["A1"] = "C.C. / NIT con DV:"
        ws["A2"] = "No. de cuenta:"
        ws["A3"] = "% participación:"

    rules = [
        MappingRule(
            field_name="cc_nit_dv",
            label="C.C. / NIT con DV:",
            target_strategy="offset_or_infer",
            required=False,
        ),
        MappingRule(
            field_name="num_cuenta",
            label="No. de cuenta:",
            target_strategy="offset_or_infer",
            required=False,
        ),
        MappingRule(
            field_name="porcentaje",
            label="% participación:",
            target_strategy="offset_or_infer",
            required=False,
        ),
    ]
    data = {
        "cc_nit_dv": "12345678-9",
        "num_cuenta": "001-234567",
        "porcentaje": "50%",
    }
    expected = {"B1": "12345678-9", "B2": "001-234567", "B3": "50%"}
    return _make_wb_bytes(setup), rules, data, expected


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------

CASES = [
    ("Layout A – label col B, valor col C",      layout_a),
    ("Layout B – label col A, valor col B",      layout_b),
    ("Layout C – label arriba, valor abajo",     layout_c),
    ("Layout D – celda vacía 2 cols derecha",    layout_d),
    ("Layout E – múltiples labels similares",    layout_e),
    ("Layout F – merged header, valor en campo", layout_f),
    ("Layout G – label NIT/CC: con barra",       layout_g),
    ("P3 – offset=0,0 no sobreescribe label",   p3_offset_zero_infers_correctly),
    ("P4 – no sobreescribe celda ocupada",       p4_no_overwrite_occupied),
    ("P5 – labels con chars especiales",         p5_special_labels),
]


def main() -> int:
    print("=" * 70)
    print("FormBot — Excel mapping experimental tests")
    print("=" * 70)

    passed = 0
    failed = 0

    for case_name, case_fn in CASES:
        try:
            wb_bytes, rules, data, expected = case_fn()
        except Exception as exc:
            print(f"\n[SETUP ERROR] {case_name}")
            traceback.print_exc()
            failed += 1
            continue

        ok, detail = _run_case(
            name=case_name.replace(" ", "_").replace("–", "-"),
            wb_bytes=wb_bytes,
            rules=rules,
            data=data,
            expected=expected,
        )
        status = "PASS" if ok else "FAIL"
        print(f"\n[{status}] {case_name}")
        if not ok:
            print(f"       {detail}")
        else:
            passed += 1
            failed += 0  # already handled

        if not ok:
            failed += 1

    print("\n" + "=" * 70)
    print(f"Results: {passed} passed, {failed} failed out of {len(CASES)} cases")
    print("=" * 70)
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
