"""Regresión Phase 4 — Section-aware scanning & hint_row disambiguation.

Verifica los 7 síntomas del brief:
  S1  Scanner detecta el mismo label en secciones distintas.
  S2  Scanner NO duplica el mismo label dentro del mismo bucket de sección.
  S3  DetectedField lleva row/column de donde se detectó la etiqueta.
  S4  hint_row selecciona la instancia de label más próxima al escaneo.
  S5  Sin hint_row se mantiene el comportamiento previo (fila más baja gana).
  S6  MappingRuleError de ambigüedad NO se lanza cuando hay hint_row.
  S7  End-to-end: formulario multi-sección escribe valores en las celdas correctas.
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "src"))

openpyxl = pytest.importorskip("openpyxl", reason="openpyxl no disponible")

from openpyxl import Workbook

from formbot.domain.models import CellPosition, MappingRule


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _wb_with_labels(labels: list[tuple[int, int, str]], *, sheet: str = "Hoja1") -> Workbook:
    """Crea un workbook con labels en posiciones dadas y una celda vacía a la derecha."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for row, col, text in labels:
        ws.cell(row=row, column=col, value=text)
        # celda destino vacía a la derecha (col+1)
        ws.cell(row=row, column=col + 1, value=None)
    return wb


# ---------------------------------------------------------------------------
# S1 — Scanner detecta duplicados en secciones distintas
# ---------------------------------------------------------------------------

class TestS1ScannerDetectaDuplicadosEnSecciones:

    def test_mismo_label_en_dos_secciones_detectado_dos_veces(self, tmp_path: Path) -> None:
        from formbot.infrastructure.document_scanners.field_scanner import (
            _SECTION_BUCKET_ROWS,
            scan_document,
        )
        path = tmp_path / "multi.xlsx"
        wb = _wb_with_labels([
            (2,  1, "Nombre:"),                          # bucket 0
            (2 + _SECTION_BUCKET_ROWS, 1, "Nombre:"),   # bucket 1
        ])
        wb.save(path)
        wb.close()

        fields = scan_document(path)
        labels = [f.label for f in fields]
        assert labels.count("Nombre:") == 2, (
            f"Se esperaban 2 instancias de 'Nombre:' pero se encontraron {labels.count('Nombre:')}"
        )

    def test_tres_secciones_tres_detecciones(self, tmp_path: Path) -> None:
        from formbot.infrastructure.document_scanners.field_scanner import (
            _SECTION_BUCKET_ROWS,
            scan_document,
        )
        path = tmp_path / "triple.xlsx"
        wb = _wb_with_labels([
            (2,                           1, "Ciudad:"),
            (2 + _SECTION_BUCKET_ROWS,    1, "Ciudad:"),
            (2 + _SECTION_BUCKET_ROWS * 2, 1, "Ciudad:"),
        ])
        wb.save(path)
        wb.close()

        fields = scan_document(path)
        assert sum(1 for f in fields if f.label == "Ciudad:") == 3


# ---------------------------------------------------------------------------
# S2 — Deduplicación dentro del mismo bucket (comportamiento conservado)
# ---------------------------------------------------------------------------

class TestS2DeduplicacionDentroDelMismoBucket:

    def test_mismo_label_en_filas_cercanas_detectado_una_vez(self, tmp_path: Path) -> None:
        from formbot.infrastructure.document_scanners.field_scanner import (
            _SECTION_BUCKET_ROWS,
            scan_document,
        )
        path = tmp_path / "dedup.xlsx"
        # Ambas filas en bucket 0 (row // 20 == 0 para row 2 y row 5)
        row1 = 2
        row2 = row1 + 3                   # mismo bucket
        assert row1 // _SECTION_BUCKET_ROWS == row2 // _SECTION_BUCKET_ROWS
        wb = _wb_with_labels([(row1, 1, "NIT:"), (row2, 1, "NIT:")])
        wb.save(path)
        wb.close()

        fields = scan_document(path)
        assert sum(1 for f in fields if f.label == "NIT:") == 1


# ---------------------------------------------------------------------------
# S3 — DetectedField lleva row y column
# ---------------------------------------------------------------------------

class TestS3DetectedFieldLlevaPosition:

    def test_row_y_column_se_populan(self, tmp_path: Path) -> None:
        from formbot.infrastructure.document_scanners.field_scanner import scan_document

        path = tmp_path / "pos.xlsx"
        wb = _wb_with_labels([(7, 3, "Correo:")])
        wb.save(path)
        wb.close()

        fields = scan_document(path)
        match = next((f for f in fields if f.label == "Correo:"), None)
        assert match is not None
        assert match.row == 7
        assert match.column == 3

    def test_sheet_se_popula(self, tmp_path: Path) -> None:
        from formbot.infrastructure.document_scanners.field_scanner import scan_document

        path = tmp_path / "sheetname.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        ws.cell(row=4, column=2, value="NIT:")
        ws.cell(row=4, column=3, value=None)
        wb.save(path)
        wb.close()

        fields = scan_document(path)
        match = next((f for f in fields if f.label == "NIT:"), None)
        assert match is not None
        assert match.sheet == "Datos"


# ---------------------------------------------------------------------------
# S4 — hint_row selecciona el candidato más cercano
# ---------------------------------------------------------------------------

class TestS4HintRowSeleccionaCandidatoCercano:

    def _build_template(self, tmp_path: Path) -> Path:
        """Label 'Nombre:' en filas 5 y 30; valor vacío a la derecha."""
        path = tmp_path / "multi_nombre.xlsx"
        wb = _wb_with_labels([(5, 1, "Nombre:"), (30, 1, "Nombre:")])
        wb.save(path)
        wb.close()
        return path

    def test_hint_row_cercano_a_5_elige_fila_5(self, tmp_path: Path) -> None:
        from formbot.application.precision_fill import PrecisionFillUseCase

        template = self._build_template(tmp_path)
        rule = MappingRule(
            field_name="nombre_empresa",
            label="Nombre:",
            target_strategy="offset_or_infer",
            required=False,
            hint_row=5,
        )
        out = tmp_path / "out5.xlsx"
        uc = PrecisionFillUseCase(template, strict_mode=False, min_confidence=0.4)
        try:
            result = uc.execute({"nombre_empresa": "Empresa S.A."}, [rule], out)
        finally:
            uc.close()

        wb = openpyxl.load_workbook(out)
        try:
            assert wb.active.cell(row=5, column=2).value == "Empresa S.A."
            assert wb.active.cell(row=30, column=2).value is None
        finally:
            wb.close()

    def test_hint_row_cercano_a_30_elige_fila_30(self, tmp_path: Path) -> None:
        from formbot.application.precision_fill import PrecisionFillUseCase

        template = self._build_template(tmp_path)
        rule = MappingRule(
            field_name="nombre_rl",
            label="Nombre:",
            target_strategy="offset_or_infer",
            required=False,
            hint_row=30,
        )
        out = tmp_path / "out30.xlsx"
        uc = PrecisionFillUseCase(template, strict_mode=False, min_confidence=0.4)
        try:
            result = uc.execute({"nombre_rl": "Juan Pérez"}, [rule], out)
        finally:
            uc.close()

        wb = openpyxl.load_workbook(out)
        try:
            assert wb.active.cell(row=30, column=2).value == "Juan Pérez"
            assert wb.active.cell(row=5, column=2).value is None
        finally:
            wb.close()


# ---------------------------------------------------------------------------
# S5 — Sin hint_row, etiqueta única se escribe normalmente
# ---------------------------------------------------------------------------

class TestS5SinHintRowEtiquetaUnica:

    def test_sin_hint_row_etiqueta_unica_escribe(self, tmp_path: Path) -> None:
        """Sin hint_row, una etiqueta única sigue funcionando sin errores."""
        from formbot.application.precision_fill import PrecisionFillUseCase

        path = tmp_path / "nohint_unica.xlsx"
        wb = _wb_with_labels([(5, 1, "Ciudad:")])
        wb.save(path)
        wb.close()

        rule = MappingRule(
            field_name="ciudad",
            label="Ciudad:",
            target_strategy="offset_or_infer",
            required=False,
        )
        out = tmp_path / "out_nohint.xlsx"
        uc = PrecisionFillUseCase(path, strict_mode=False, min_confidence=0.4)
        try:
            uc.execute({"ciudad": "Bogotá"}, [rule], out)
        finally:
            uc.close()

        wb2 = openpyxl.load_workbook(out)
        try:
            assert wb2.active.cell(row=5, column=2).value == "Bogotá"
        finally:
            wb2.close()

    def test_sin_hint_row_etiqueta_duplicada_bloquea_campo(self, tmp_path: Path) -> None:
        """Sin hint_row, etiqueta duplicada con puntuaciones iguales se bloquea (seguridad).
        El sistema no escribe en la celda equivocada — prefiere no escribir a escribir mal.
        """
        from formbot.application.precision_fill import PrecisionFillUseCase

        path = tmp_path / "nohint_dup.xlsx"
        wb = _wb_with_labels([(5, 1, "Ciudad:"), (25, 1, "Ciudad:")])
        wb.save(path)
        wb.close()

        rule = MappingRule(
            field_name="ciudad",
            label="Ciudad:",
            target_strategy="offset_or_infer",
            required=False,
        )
        out = tmp_path / "out_blocked.xlsx"
        uc = PrecisionFillUseCase(path, strict_mode=False, min_confidence=0.4)
        try:
            result = uc.execute({"ciudad": "Bogotá"}, [rule], out)
        finally:
            uc.close()

        # Campo bloqueado: ambigüedad sin hint_row es el comportamiento correcto (seguro)
        assert "ciudad" in result.blocked_fields


# ---------------------------------------------------------------------------
# S6 — Sin hint_row lanza ambigüedad; con hint_row NO la lanza
# ---------------------------------------------------------------------------

class TestS6AmbiguedadSuprimidaConHintRow:

    def _build_two_label_wb(self, tmp_path: Path) -> Path:
        path = tmp_path / "ambig.xlsx"
        wb = _wb_with_labels([(5, 1, "Nombre:"), (30, 1, "Nombre:")])
        wb.save(path)
        wb.close()
        return path

    def test_con_hint_row_no_lanza_mapping_rule_error(self, tmp_path: Path) -> None:
        from formbot.application.precision_fill import PrecisionFillUseCase

        template = self._build_two_label_wb(tmp_path)
        rule = MappingRule(
            field_name="nombre",
            label="Nombre:",
            target_strategy="offset_or_infer",
            required=False,
            hint_row=5,
        )
        out = tmp_path / "no_error.xlsx"
        uc = PrecisionFillUseCase(template, strict_mode=False, min_confidence=0.4)
        try:
            # No debe lanzar excepción
            uc.execute({"nombre": "Test"}, [rule], out)
        finally:
            uc.close()


# ---------------------------------------------------------------------------
# S7 — End-to-end: formulario multi-sección escribe en celdas correctas
# ---------------------------------------------------------------------------

class TestS7EndToEndMultiSeccion:

    def test_dos_secciones_con_mismo_label_escriben_separado(self, tmp_path: Path) -> None:
        """Simula CU-FOR-001: 'Nombre:' en sección empresa (fila 5) y sección RL (fila 30).
        Cada regla tiene hint_row → cada valor va a la celda correcta.
        """
        from formbot.application.precision_fill import PrecisionFillUseCase

        template = tmp_path / "form_multi.xlsx"
        wb = _wb_with_labels([
            (5,  1, "Nombre:"),   # Sección: empresa
            (30, 1, "Nombre:"),   # Sección: representante legal
        ])
        wb.save(template)
        wb.close()

        rules = [
            MappingRule(
                field_name="nombre_empresa",
                label="Nombre:",
                target_strategy="offset_or_infer",
                required=False,
                hint_row=5,
            ),
            MappingRule(
                field_name="nombre_rl",
                label="Nombre:",
                target_strategy="offset_or_infer",
                required=False,
                hint_row=30,
            ),
        ]
        data = {
            "nombre_empresa": "ASTECO SAS",
            "nombre_rl": "Carlos García",
        }
        out = tmp_path / "filled_multi.xlsx"
        uc = PrecisionFillUseCase(template, strict_mode=False, min_confidence=0.4)
        try:
            result = uc.execute(data, rules, out)
        finally:
            uc.close()

        wb2 = openpyxl.load_workbook(out)
        try:
            ws = wb2.active
            assert ws.cell(row=5,  column=2).value == "ASTECO SAS",   "Empresa incorrecta"
            assert ws.cell(row=30, column=2).value == "Carlos García", "RL incorrecto"
        finally:
            wb2.close()
