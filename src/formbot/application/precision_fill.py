from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
import logging
from pathlib import Path
import re
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.workbook.workbook import Workbook

from formbot.domain.exceptions import (
    DataValidationError,
    DocumentProcessingError,
    DocumentSaveError,
    MappingRuleError,
    PositionOutOfBoundsError,
    ValidationException,
)
from formbot.domain.models import CellPosition, MappingRule
from formbot.infrastructure.document_readers.excel_structure_reader import (
    ExcelStructureReader,
    LabelCandidate,
)
from formbot.infrastructure.document_writers.excel_value_writer import ExcelValueWriter
from formbot.shared.utils import find_duplicates, normalize_text

LOGGER = logging.getLogger(__name__)

# Constantes de inferencia de celda destino
_INFER_RIGHT_SCAN: int = 7     # columnas a escanear hacia la derecha
_INFER_DOWN_SCAN: int = 3      # filas a escanear hacia abajo
_INFER_DOWN_PENALTY: float = 0.08   # penalidad por preferir abajo vs derecha
_INFER_MIN_SCORE: float = 0.55      # score minimo para aceptar inferencia


@dataclass(frozen=True)
class PrecisionDecision:
    field_name: str
    status: str
    confidence: float
    reason: str
    label: str | None = None
    label_position: CellPosition | None = None
    target_position: CellPosition | None = None
    value_preview: str | None = None


@dataclass(frozen=True)
class PrecisionFillResult:
    output_path: Path
    strict_mode: bool
    min_confidence: float
    decisions: list[PrecisionDecision] = field(default_factory=list)
    blocked_fields: list[str] = field(default_factory=list)
    written_fields: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class _Resolution:
    confidence: float
    reason: str
    label_text: str
    label_position: CellPosition
    target_position: CellPosition
    write_value: Any


class PrecisionFillUseCase:
    def __init__(
        self,
        template_path: Path,
        *,
        strict_mode: bool = True,
        min_confidence: float = 0.85,
        allow_overwrite_existing: bool = False,
        infer_right_scan: int = _INFER_RIGHT_SCAN,
        infer_down_scan: int = _INFER_DOWN_SCAN,
    ) -> None:
        self._strict_mode = strict_mode
        self._min_confidence = min_confidence
        self._allow_overwrite_existing = allow_overwrite_existing
        self._infer_right_scan = infer_right_scan
        self._infer_down_scan = infer_down_scan
        self._structure_reader = ExcelStructureReader()
        self._value_writer = ExcelValueWriter()
        self._workbook = self._load_template(template_path)

    def execute(
        self,
        data: Mapping[str, Any],
        mapping_rules: Sequence[MappingRule],
        output_path: Path,
    ) -> PrecisionFillResult:
        self._validate_input(data, mapping_rules)

        decisions: list[PrecisionDecision] = []
        ready_to_write: list[tuple[MappingRule, _Resolution]] = []

        LOGGER.info(
            "Inicio precision_fill: %d reglas, min_confidence=%.2f, strict=%s",
            len(mapping_rules),
            self._min_confidence,
            self._strict_mode,
        )

        for rule in mapping_rules:
            field_present = rule.field_name in data
            if not field_present:
                if rule.required:
                    raise DataValidationError(
                        f"Falta campo requerido en payload: {rule.field_name}"
                    )
                LOGGER.info("Campo opcional omitido (sin datos): '%s'", rule.field_name)
                decisions.append(
                    PrecisionDecision(
                        field_name=rule.field_name,
                        status="skipped",
                        confidence=1.0,
                        reason="Campo opcional no recibido en payload",
                    )
                )
                continue

            value = data[rule.field_name]
            try:
                resolution = self._resolve_field(rule, value)
            except (
                MappingRuleError,
                PositionOutOfBoundsError,
                DataValidationError,
                ValidationException,
            ) as exc:
                LOGGER.warning("Campo '%s' bloqueado por error: %s", rule.field_name, exc)
                decisions.append(
                    PrecisionDecision(
                        field_name=rule.field_name,
                        status="blocked",
                        confidence=0.0,
                        reason=str(exc),
                        value_preview=_preview(value),
                    )
                )
                continue
            threshold = (
                rule.confidence_threshold
                if rule.confidence_threshold is not None
                else self._min_confidence
            )
            if resolution.confidence < threshold:
                LOGGER.warning(
                    "Campo '%s' bloqueado por baja confianza: %.4f < %.2f. %s",
                    rule.field_name,
                    resolution.confidence,
                    threshold,
                    resolution.reason,
                )
                decisions.append(
                    PrecisionDecision(
                        field_name=rule.field_name,
                        status="blocked",
                        confidence=resolution.confidence,
                        reason=(
                            f"{resolution.reason}. "
                            f"Confianza {resolution.confidence:.2f} < umbral {threshold:.2f}"
                        ),
                        label=resolution.label_text,
                        label_position=resolution.label_position,
                        target_position=resolution.target_position,
                        value_preview=_preview(value),
                    )
                )
                continue

            LOGGER.debug(
                "Campo '%s' listo: confianza=%.4f, label='%s', destino=%s!R%dC%d",
                rule.field_name,
                resolution.confidence,
                resolution.label_text,
                resolution.target_position.sheet_name,
                resolution.target_position.row,
                resolution.target_position.column,
            )
            decisions.append(
                PrecisionDecision(
                    field_name=rule.field_name,
                    status="ready",
                    confidence=resolution.confidence,
                    reason=resolution.reason,
                    label=resolution.label_text,
                    label_position=resolution.label_position,
                    target_position=resolution.target_position,
                    value_preview=_preview(value),
                )
            )
            ready_to_write.append((rule, resolution))

        blocked_fields = [item.field_name for item in decisions if item.status == "blocked"]
        if self._strict_mode and blocked_fields:
            LOGGER.error(
                "Modo estricto: se cancela escritura por %d campo(s) bloqueado(s): %s",
                len(blocked_fields),
                ", ".join(blocked_fields),
            )
            raise DataValidationError(
                "Modo estricto bloqueo de escritura por campos sin confianza suficiente: "
                + ", ".join(blocked_fields)
            )

        written_fields: list[str] = []
        for rule, resolution in ready_to_write:
            try:
                self._value_writer.write_value(
                    self._workbook,
                    resolution.target_position,
                    resolution.write_value,
                )
            except (PositionOutOfBoundsError, ValidationException, MappingRuleError):
                raise
            except Exception as exc:  # pragma: no cover - defensa.
                raise DocumentSaveError(
                    f"Fallo al escribir campo '{rule.field_name}' en "
                    f"{resolution.target_position.sheet_name}!R{resolution.target_position.row}C{resolution.target_position.column}"
                ) from exc
            LOGGER.info("Campo '%s' escrito correctamente.", rule.field_name)
            written_fields.append(rule.field_name)

        self._value_writer.save(self._workbook, output_path)
        LOGGER.info(
            "precision_fill finalizado: escritos=%d, bloqueados=%d, omitidos=%d. Salida: %s",
            len(written_fields),
            len(blocked_fields),
            sum(1 for d in decisions if d.status == "skipped"),
            output_path,
        )
        final_decisions = [
            decision
            if decision.status != "ready"
            else PrecisionDecision(
                field_name=decision.field_name,
                status="written",
                confidence=decision.confidence,
                reason=decision.reason,
                label=decision.label,
                label_position=decision.label_position,
                target_position=decision.target_position,
                value_preview=decision.value_preview,
            )
            for decision in decisions
        ]
        return PrecisionFillResult(
            output_path=output_path,
            strict_mode=self._strict_mode,
            min_confidence=self._min_confidence,
            decisions=final_decisions,
            blocked_fields=blocked_fields,
            written_fields=written_fields,
        )

    def close(self) -> None:
        self._workbook.close()

    @staticmethod
    def _load_template(template_path: Path) -> Workbook:
        if not template_path.exists():
            raise DocumentProcessingError(
                f"No existe el template de Excel: {template_path}"
            )
        if template_path.suffix.lower() == ".xls":
            raise DocumentProcessingError(
                f"El formato .xls (Excel 97-2003) no está soportado. "
                f"Convierta el archivo a .xlsx o .xlsm antes de usarlo: {template_path}"
            )
        try:
            return load_workbook(template_path)
        except Exception as exc:  # pragma: no cover - depende del archivo.
            raise DocumentProcessingError(
                f"No fue posible cargar el template de Excel: {template_path}"
            ) from exc

    @staticmethod
    def _validate_input(data: Mapping[str, Any], mapping_rules: Sequence[MappingRule]) -> None:
        if not mapping_rules:
            raise MappingRuleError("No hay reglas de mapeo configuradas")
        duplicated = find_duplicates([rule.field_name for rule in mapping_rules])
        if duplicated:
            raise MappingRuleError(
                "Campos duplicados en mapping: " + ", ".join(sorted(duplicated))
            )
        if not isinstance(data, Mapping):
            raise DataValidationError("El payload de datos debe ser tipo objeto/diccionario")

    def _resolve_field(self, rule: MappingRule, value: Any) -> _Resolution:
        label_candidate, label_score = self._select_label_candidate(rule)
        target_position, target_score, reason = self._resolve_target(
            rule,
            label_candidate.position,
            value,
        )
        self._validate_value_type(rule, value)
        write_value = self._resolve_write_value(rule, value)
        confidence = round(
            max(0.0, min(1.0, (label_score * 0.65) + (target_score * 0.35))),
            4,
        )
        return _Resolution(
            confidence=confidence,
            reason=reason,
            label_text=label_candidate.text,
            label_position=label_candidate.position,
            target_position=target_position,
            write_value=write_value,
        )

    def _select_label_candidate(self, rule: MappingRule) -> tuple[LabelCandidate, float]:
        search_terms = [rule.label, *rule.aliases]
        weighted: list[tuple[LabelCandidate, float, float]] = []
        for idx, term in enumerate(search_terms):
            candidates = self._structure_reader.find_label_candidates(
                workbook=self._workbook,
                text=term,
                sheet_name=rule.sheet_name,
            )
            normalized_term = normalize_text(term)
            if len(normalized_term) <= 3:
                candidates = [c for c in candidates if c.match_type == "exact"]
            for candidate in candidates:
                base_score = _score_label(candidate, term, idx)
                context_adjustment = self._candidate_context_adjustment(rule, candidate)
                final_score = max(0.0, min(1.0, base_score + context_adjustment))
                weighted.append((candidate, final_score, base_score))

        if not weighted:
            raise MappingRuleError(
                f"No se encontro etiqueta para campo '{rule.field_name}' "
                f"(label='{rule.label}', aliases={list(rule.aliases)})"
            )

        # Si se conoce la fila esperada (hint_row), ordenar por proximidad a esa fila
        # para seleccionar la instancia correcta cuando el label se repite en múltiples secciones.
        if rule.hint_row is not None:
            hint = rule.hint_row
            weighted.sort(
                key=lambda item: (
                    item[1],
                    item[2],
                    -abs(item[0].position.row - hint),
                ),
                reverse=True,
            )
        else:
            weighted.sort(
                key=lambda item: (
                    item[1],
                    item[2],
                    -item[0].position.row,
                    -item[0].position.column,
                ),
                reverse=True,
            )
        best_candidate, best_score, _ = weighted[0]
        if len(weighted) == 1:
            return best_candidate, best_score

        second_candidate, second_score, _ = weighted[1]
        # No lanzar error de ambigüedad cuando hay hint_row: la proximidad ya desambiguó.
        if (
            rule.hint_row is None
            and abs(best_score - second_score) < 0.05
            and (
                best_candidate.position.sheet_name != second_candidate.position.sheet_name
                or best_candidate.position.row != second_candidate.position.row
                or best_candidate.position.column != second_candidate.position.column
            )
        ):
            raise MappingRuleError(
                f"Etiqueta ambigua para campo '{rule.field_name}'. "
                f"Top candidatos con score cercano: "
                f"{_fmt_candidate(best_candidate, best_score)} | {_fmt_candidate(second_candidate, second_score)}"
            )
        return best_candidate, best_score

    def _candidate_context_adjustment(
        self,
        rule: MappingRule,
        candidate: LabelCandidate,
    ) -> float:
        field_name = rule.field_name
        label_text = normalize_text(candidate.text)
        row_context = self._row_context_text(candidate.position)
        same_row_context = self._row_context_text(candidate.position, rows_back=0)

        adjustment = 0.0
        if field_name in {"numero_identificacion_nit", "nit"}:
            if any(
                token in row_context
                for token in ("gravado", "exento", "comercial", "transporte")
            ):
                adjustment -= 0.14
            if any(token in row_context for token in ("identificacion", "numero")):
                adjustment += 0.10
            if "nit" in label_text:
                adjustment += 0.02
        elif field_name in {"representante_legal_nombre", "nombre_representante"}:
            if "firma" in label_text:
                adjustment -= 0.18
            if any(
                token in label_text
                for token in ("nombre representante legal", "nombres y apellidos")
            ):
                adjustment += 0.08
        elif field_name in {"ciudad_municipio", "ciudad"}:
            if any(token in label_text for token in ("fecha", "solicitud")):
                adjustment -= 0.12
            if "swift" in row_context:
                adjustment -= 0.14
            if label_text in {"ciudad", "ciudad:"}:
                adjustment += 0.08
        elif field_name == "pais":
            if any(token in row_context for token in ("swift", "iban", "aba")):
                adjustment -= 0.14
            if label_text in {"pais", "pais:"}:
                adjustment += 0.06
        elif field_name in {"telefono_fijo", "telefono"}:
            if any(token in row_context for token in ("movil", "celular")):
                adjustment -= 0.08
            if any(token in label_text for token in ("telefonos", "telefono")):
                adjustment += 0.04
        elif field_name == "celular":
            if "contacto pagos" in same_row_context:
                adjustment -= 0.08
            if "contacto comercial" in same_row_context:
                adjustment += 0.03
        elif field_name in {"correo_electronico", "correo", "email"}:
            if any(
                token in row_context
                for token in ("contacto comercial", "contacto pagos", "cartera")
            ):
                adjustment -= 0.09
            if any(token in row_context for token in ("direccion", "telefonos", "telefono")):
                adjustment += 0.04
        elif field_name in {"banco_nombre", "tipo_cuenta", "numero_cuenta"}:
            if "exterior" in row_context:
                adjustment -= 0.15
            if "nacional" in row_context:
                adjustment += 0.10

        return max(-0.25, min(0.20, adjustment))

    def _row_context_text(self, position: CellPosition, *, rows_back: int = 2) -> str:
        sheet = self._workbook[position.sheet_name]
        min_col = 1
        max_col = min(sheet.max_column, position.column + 20)
        texts: list[str] = []
        min_row = max(1, position.row - rows_back)
        for row in range(min_row, position.row + 1):
            for column in range(min_col, max_col + 1):
                value = sheet.cell(row=row, column=column).value
                if not isinstance(value, str):
                    continue
                if not value.strip():
                    continue
                texts.append(value)
        return normalize_text(" ".join(texts))

    def _resolve_write_value(self, rule: MappingRule, value: Any) -> Any:
        if rule.write_mode == "mark":
            return rule.mark_symbol
        return value

    def _resolve_mark_target(
        self,
        rule: MappingRule,
        label_position: CellPosition,
        value: Any,
    ) -> tuple[CellPosition, float, str] | None:
        selection = _normalize_mark_selection(value)
        if selection is None:
            raise DataValidationError(
                f"Campo '{rule.field_name}' requiere seleccion valida para marcacion "
                "(si/no/true/false/x/check)"
            )

        if rule.row_offset != 0 or rule.column_offset != 0:
            target = label_position.shifted(rule.row_offset, rule.column_offset)
            score = self._target_quality_score(target)
            return target, score, "Destino de marcacion por offset explicito"

        sheet = self._workbook[label_position.sheet_name]
        option_cell = self._find_option_cell(
            sheet_name=label_position.sheet_name,
            row=label_position.row,
            start_col=max(1, label_position.column - 1),
            selection=selection,
        )
        if option_cell is not None:
            option_row, option_col, option_text = option_cell
            for candidate_col in _mark_candidate_columns(option_col, option_text, selection):
                if candidate_col < 1:
                    continue
                candidate = CellPosition(
                    sheet_name=label_position.sheet_name,
                    row=option_row,
                    column=candidate_col,
                )
                score = self._target_quality_score(candidate)
                if score >= _INFER_MIN_SCORE:
                    return candidate, score, "Destino de marcacion por opcion detectada"

            # Fallback local: busca alrededor de la celda de opcion detectada.
            local_best: tuple[CellPosition, float] | None = None
            min_local_col = max(1, option_col - 2)
            max_local_col = min(sheet.max_column, option_col + 2)
            for col in range(min_local_col, max_local_col + 1):
                if col == option_col:
                    continue
                candidate = CellPosition(
                    sheet_name=label_position.sheet_name,
                    row=option_row,
                    column=col,
                )
                score = self._target_quality_score(candidate)
                if local_best is None or score > local_best[1]:
                    local_best = (candidate, score)
            if local_best is not None and local_best[1] >= _INFER_MIN_SCORE:
                return (
                    local_best[0],
                    local_best[1],
                    "Destino de marcacion por vecindad de opcion",
                )
            if selection not in {"si", "no"}:
                raise DataValidationError(
                    f"No se detecto cuadro editable para opcion '{selection}' en campo '{rule.field_name}'"
                )
        elif selection not in {"si", "no"}:
            raise DataValidationError(
                f"No se encontro opcion '{selection}' para campo '{rule.field_name}'"
            )

        # Fallback global: misma inferencia general (derecha/abajo) del motor.
        inferred = self._infer_target(rule, label_position)
        if inferred is None:
            return None
        target, target_score = inferred
        return target, target_score, "Destino de marcacion por inferencia general"

    def _find_option_cell(
        self,
        *,
        sheet_name: str,
        row: int,
        start_col: int,
        selection: str,
    ) -> tuple[int, int, str] | None:
        sheet = self._workbook[sheet_name]
        max_col = min(sheet.max_column, start_col + 12)
        min_row = max(1, row - 1)
        max_row = min(sheet.max_row, row + 1)
        best: tuple[int, int, str, float] | None = None

        for candidate_row in range(min_row, max_row + 1):
            for col in range(start_col, max_col + 1):
                raw_value = sheet.cell(row=candidate_row, column=col).value
                if not isinstance(raw_value, str):
                    continue
                option_text = normalize_text(raw_value)
                if not option_text:
                    continue

                has_yes = _contains_yes_token(option_text)
                has_no = _contains_no_token(option_text)
                score = 0.0

                if selection == "si":
                    if has_yes:
                        score += 0.88
                    if has_yes and has_no:
                        score += 0.06
                elif selection == "no":
                    if has_no:
                        score += 0.88
                    if has_yes and has_no:
                        score += 0.06
                elif selection in option_text:
                    score += 0.9

                if score <= 0:
                    continue

                col_penalty = 0.01 * max(0, col - start_col)
                row_penalty = 0.07 * abs(candidate_row - row)
                score -= (col_penalty + row_penalty)

                if best is None or score > best[3]:
                    best = (candidate_row, col, option_text, score)

        if best is None:
            return None
        return best[0], best[1], best[2]

    def _resolve_target(
        self,
        rule: MappingRule,
        label_position: CellPosition,
        value: Any,
    ) -> tuple[CellPosition, float, str]:
        if rule.write_mode == "mark":
            mark_resolution = self._resolve_mark_target(rule, label_position, value)
            if mark_resolution is None:
                raise MappingRuleError(
                    f"No se pudo inferir celda de marcacion para campo '{rule.field_name}' "
                    f"desde label en {label_position.sheet_name}!R{label_position.row}C{label_position.column}"
                )
            target, score, reason = mark_resolution
            return target, score, reason

        if rule.target_strategy == "offset":
            target = label_position.shifted(rule.row_offset, rule.column_offset)
            score = self._target_quality_score(target)
            return target, score, "Destino por offset explicito"

        if rule.target_strategy == "infer":
            anchored = self._infer_target_from_number_anchor(rule, label_position)
            if anchored is not None:
                target, target_score = anchored
                return target, target_score, "Destino inferido por ancla numerica"
            inferred = self._infer_target(rule, label_position)
            if inferred is None:
                raise MappingRuleError(
                    f"No se pudo inferir celda destino para campo '{rule.field_name}' "
                    f"desde label en {label_position.sheet_name}!R{label_position.row}C{label_position.column}"
                )
            target, target_score = inferred
            return target, target_score, "Destino inferido por vecindad"

        # offset_or_infer
        if rule.row_offset != 0 or rule.column_offset != 0:
            target = label_position.shifted(rule.row_offset, rule.column_offset)
            score = self._target_quality_score(target)
            return target, score, "Destino por offset explicito"

        anchored = self._infer_target_from_number_anchor(rule, label_position)
        if anchored is not None:
            target, target_score = anchored
            return target, target_score, "Destino inferido por ancla numerica"

        inferred = self._infer_target(rule, label_position)
        if inferred is None:
            raise MappingRuleError(
                f"No se pudo inferir celda destino para campo '{rule.field_name}' "
                f"y no tiene offset configurado"
            )
        target, target_score = inferred
        return target, target_score, "Destino inferido por vecindad"

    def _infer_target_from_number_anchor(
        self,
        rule: MappingRule,
        label_position: CellPosition,
    ) -> tuple[CellPosition, float] | None:
        if rule.value_type not in {"number", "numeric", "nit"}:
            return None

        sheet = self._workbook[label_position.sheet_name]
        anchor_col = None
        for col in range(label_position.column + 1, min(sheet.max_column, label_position.column + 20) + 1):
            cell_value = sheet.cell(row=label_position.row, column=col).value
            if not isinstance(cell_value, str):
                continue
            normalized = normalize_text(cell_value)
            if any(token in normalized for token in ("numero", "identificacion", "documento", "no")):
                anchor_col = col
                break

        if anchor_col is None:
            label_text = normalize_text(str(sheet.cell(row=label_position.row, column=label_position.column).value or ""))
            if any(token in label_text for token in ("numero", "identificacion", "documento", "cuenta", "nit", "no")):
                anchor_col = label_position.column

        if anchor_col is None:
            return None

        best: tuple[CellPosition, float] | None = None
        for col in range(anchor_col + 1, min(sheet.max_column, anchor_col + 8) + 1):
            position = CellPosition(
                sheet_name=label_position.sheet_name,
                row=label_position.row,
                column=col,
            )
            score = self._target_quality_score(position)
            if score <= 0:
                continue

            score += self._typed_existing_bonus(position, rule.value_type)
            ranking = (score, -abs(col - anchor_col))
            if best is None or ranking > (best[1], -abs(best[0].column - anchor_col)):
                best = (position, score)

        if best is None or best[1] < _INFER_MIN_SCORE:
            return None
        return best[0], max(0.0, min(1.0, best[1]))

    def _infer_target(self, rule: MappingRule, label_position: CellPosition) -> tuple[CellPosition, float] | None:
        sheet = self._workbook[label_position.sheet_name]
        best: tuple[CellPosition, float] | None = None

        # Prioridad: derecha en misma fila.
        for delta in range(1, self._infer_right_scan + 1):
            candidate = CellPosition(
                sheet_name=label_position.sheet_name,
                row=label_position.row,
                column=label_position.column + delta,
            )
            score = self._target_quality_score(candidate) + self._typed_existing_bonus(
                candidate,
                rule.value_type,
            )
            if best is None or score > best[1]:
                best = (candidate, score)

        # Segunda prioridad: hacia abajo en misma columna.
        for delta in range(1, self._infer_down_scan + 1):
            candidate = CellPosition(
                sheet_name=label_position.sheet_name,
                row=label_position.row + delta,
                column=label_position.column,
            )
            score = (
                self._target_quality_score(candidate)
                + self._typed_existing_bonus(candidate, rule.value_type)
                - _INFER_DOWN_PENALTY
            )
            if best is None or score > best[1]:
                best = (candidate, score)

        if best is None or best[1] < _INFER_MIN_SCORE:
            return None
        # Evita escribir sobre la etiqueta.
        if best[0].row == label_position.row and best[0].column == label_position.column:
            return None
        if best[0].column > sheet.max_column + 20:
            return None
        return best[0], max(0.0, min(1.0, best[1]))

    def _typed_existing_bonus(self, position: CellPosition, value_type: str | None) -> float:
        if value_type is None:
            return 0.0
        if not self._allow_overwrite_existing:
            return 0.0
        sheet = self._workbook[position.sheet_name]
        value = sheet.cell(row=position.row, column=position.column).value
        if _matches_existing_value_type(value, value_type):
            return 0.35
        return 0.0

    def _target_quality_score(self, position: CellPosition) -> float:
        sheet = self._workbook[position.sheet_name]
        cell = sheet.cell(row=position.row, column=position.column)
        if isinstance(cell, MergedCell):
            return 0.0

        value = cell.value
        if value is None:
            return 1.0

        as_text = str(value).strip()
        if not as_text:
            return 0.98

        normalized = normalize_text(as_text)
        if _is_placeholder(normalized):
            return 0.9

        if _looks_like_label(as_text, normalized):
            return 0.35

        if self._allow_overwrite_existing:
            return 0.72
        return 0.42

    @staticmethod
    def _validate_value_type(rule: MappingRule, value: Any) -> None:
        if rule.write_mode == "mark":
            selection = _normalize_mark_selection(value)
            if selection is None:
                raise DataValidationError(
                    f"Campo '{rule.field_name}' requiere valor valido para marcacion"
                )
            return

        if rule.value_type is None:
            return

        if value is None:
            raise DataValidationError(
                f"Campo '{rule.field_name}' requiere valor para tipo '{rule.value_type}'"
            )

        value_text = str(value).strip()
        if not value_text:
            raise DataValidationError(
                f"Campo '{rule.field_name}' requiere texto no vacio para tipo '{rule.value_type}'"
            )

        type_name = rule.value_type.lower()
        if type_name == "email":
            if re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", value_text) is None:
                raise DataValidationError(
                    f"Campo '{rule.field_name}' no cumple formato email: {value_text}"
                )
            return

        if type_name in {"number", "numeric"}:
            try:
                float(value_text.replace(",", "."))
            except ValueError as exc:
                raise DataValidationError(
                    f"Campo '{rule.field_name}' no es numerico: {value_text}"
                ) from exc
            return

        if type_name == "date":
            iso_candidate = value_text.replace("/", "-")
            try:
                datetime.fromisoformat(iso_candidate)
            except ValueError as exc:
                raise DataValidationError(
                    f"Campo '{rule.field_name}' no cumple formato fecha ISO (YYYY-MM-DD): {value_text}"
                ) from exc
            return

        if type_name in {"phone", "telefono"}:
            digits = re.sub(r"\D", "", value_text)
            if len(digits) < 7:
                raise DataValidationError(
                    f"Campo '{rule.field_name}' no parece telefono valido: {value_text}"
                )
            return

        if type_name == "nit":
            digits = re.sub(r"\D", "", value_text)
            if len(digits) < 6:
                raise DataValidationError(
                    f"Campo '{rule.field_name}' no parece NIT valido: {value_text}"
                )


def _score_label(candidate: LabelCandidate, term: str, term_index: int) -> float:
    is_primary = term_index == 0
    if candidate.match_type == "exact":
        base = 1.0 if is_primary else 0.96
    else:
        base = 0.83 if is_primary else 0.78
    penalty = 0.01 * term_index
    # Leve bono por coincidir con token inicial de la etiqueta buscada.
    normalized_term = normalize_text(term)
    normalized_found = normalize_text(candidate.text)
    if normalized_term and normalized_found.startswith(normalized_term):
        base += 0.01
    return max(0.0, min(1.0, base - penalty))


def _is_placeholder(value: str) -> bool:
    if not value:
        return True
    if set(value) <= {"_", "-", ".", "x"}:
        return True
    return value in {"n/a", "na", "pendiente", "por definir"}


def _looks_like_label(raw_value: str, normalized_value: str) -> bool:
    if ":" in raw_value:
        return True
    marker_terms = (
        "nombre",
        "razon social",
        "direccion",
        "ciudad",
        "telefono",
        "correo",
        "representante",
        "firma",
        "fecha",
        "nit",
        "numero",
        "identificacion",
        "documento",
        "tipo de documento",
        "cuenta",
    )
    return any(term in normalized_value for term in marker_terms)


def _matches_existing_value_type(cell_value: Any, value_type: str | None) -> bool:
    if cell_value is None or value_type is None:
        return False
    text = str(cell_value).strip()
    if not text:
        return False
    normalized_type = value_type.lower().strip()
    digits = re.sub(r"\D", "", text)
    if normalized_type == "email":
        return re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", text) is not None
    if normalized_type in {"phone", "telefono"}:
        return len(digits) >= 7
    if normalized_type == "date":
        iso_candidate = text.replace("/", "-")
        try:
            datetime.fromisoformat(iso_candidate)
        except ValueError:
            return False
        return True
    if normalized_type == "nit":
        return len(digits) >= 6
    if normalized_type in {"number", "numeric"}:
        try:
            float(text.replace(",", "."))
        except ValueError:
            return False
        return True
    return False


def _normalize_mark_selection(value: Any) -> str | None:
    if isinstance(value, bool):
        return "si" if value else "no"

    if value is None:
        return None

    normalized = normalize_text(str(value))
    if not normalized:
        return None

    true_tokens = {
        "si",
        "yes",
        "true",
        "1",
        "x",
        "check",
        "checked",
        "seleccionado",
        "marcar",
    }
    false_tokens = {
        "no",
        "false",
        "0",
        "unchecked",
        "sin marcar",
    }

    if normalized in true_tokens:
        return "si"
    if normalized in false_tokens:
        return "no"
    return normalized


def _contains_yes_token(text: str) -> bool:
    return re.search(r"\b(si|yes)\b", text) is not None


def _contains_no_token(text: str) -> bool:
    return re.search(r"\bno\b", text) is not None


def _mark_candidate_columns(option_col: int, option_text: str, selection: str) -> list[int]:
    has_yes = _contains_yes_token(option_text)
    has_no = _contains_no_token(option_text)

    if selection == "si" and has_yes and has_no:
        return [option_col - 1, option_col - 2, option_col + 1]
    if selection == "no" and has_yes and has_no:
        return [option_col + 1, option_col + 2, option_col - 1]
    if selection == "si" and has_yes:
        return [option_col - 1, option_col + 1]
    if selection == "no" and has_no:
        return [option_col + 1, option_col - 1]
    return [option_col - 1, option_col + 1]


def _preview(value: Any) -> str:
    text = str(value)
    if len(text) <= 80:
        return text
    return text[:77] + "..."


def _fmt_candidate(candidate: LabelCandidate, score: float) -> str:
    return (
        f"{candidate.match_type}:{candidate.text!r}"
        f"@{candidate.position.sheet_name}!R{candidate.position.row}C{candidate.position.column}"
        f" score={score:.2f}"
    )


