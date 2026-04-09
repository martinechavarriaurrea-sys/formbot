"""Auto-detección de campos rellenables en documentos (Excel, PDF, Word).

El escáner analiza la estructura del documento para identificar etiquetas
(labels) que tienen celdas/campos vacíos adyacentes donde se deben escribir datos.

Retorna una lista de DetectedField con:
- field_key: clave snake_case única para usar en el payload de datos
- label: texto original de la etiqueta tal como aparece en el documento
- sheet: nombre de la hoja (Excel) o None (PDF/Word)
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from formbot.shared.utils import normalize_text

# Número máximo de columnas a escanear hacia la derecha buscando celda vacía
_INFER_RIGHT_MAX = 8
# Número máximo de filas a escanear hacia abajo buscando celda vacía
_INFER_DOWN_MAX = 3
# Longitud mínima del texto de la etiqueta
_MIN_LABEL_LEN = 2

# Palabras clave conocidas que indican que un texto corto es una etiqueta de formulario
_KNOWN_LABEL_TOKENS = {
    "nit", "nombre", "ciudad", "pais", "fecha", "correo", "email",
    "celular", "telefono", "telefax", "banco", "cargo", "firma",
    "tipo", "numero", "cuenta", "rut", "contacto", "codigo",
    "actividad", "digito", "cupo", "departamento", "dv",
    "plazo", "representante", "identificacion", "direccion",
    "patrimonio", "activos", "pasivos", "ingresos", "egresos",
    "banco", "entidad", "titular", "sucursal", "barrio",
    "sector", "ciiu", "matricula", "sigla", "objeto",
}


@dataclass
class DetectedField:
    """Campo rellenable detectado automáticamente en un documento."""

    field_key: str       # Clave snake_case (única dentro del documento)
    label: str           # Texto original de la etiqueta en el documento
    sheet: str | None    # Nombre de la hoja (Excel) o None (PDF/Word)


# ---------------------------------------------------------------------------
# Punto de entrada público
# ---------------------------------------------------------------------------

def scan_document(template_path: Path) -> list[DetectedField]:
    """Detecta automáticamente los campos rellenables en el documento."""
    suffix = template_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _scan_excel(template_path)
    if suffix == ".pdf":
        return _scan_pdf(template_path)
    if suffix == ".docx":
        return _scan_word(template_path)
    return []


def label_to_key(text: str) -> str:
    """Convierte el texto de una etiqueta a una clave snake_case (máx. 60 chars)."""
    normalized = normalize_text(text)
    return "_".join(normalized.split())[:60]


# ---------------------------------------------------------------------------
# Escáner Excel
# ---------------------------------------------------------------------------

def _scan_excel(path: Path) -> list[DetectedField]:
    try:
        from openpyxl import load_workbook
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        return []

    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return []

    fields: list[DetectedField] = []
    seen_norm: set[str] = set()

    try:
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    raw = cell.value
                    if raw is None:
                        continue
                    text = str(raw).strip()
                    if len(text) < _MIN_LABEL_LEN:
                        continue
                    if _is_numeric(text):
                        continue
                    if not _is_likely_form_label(text):
                        continue

                    norm = normalize_text(text)
                    if norm in seen_norm:
                        continue

                    if not _has_adjacent_empty_excel(sheet, cell.row, cell.column):
                        continue

                    seen_norm.add(norm)
                    fields.append(DetectedField(
                        field_key=label_to_key(text),
                        label=text,
                        sheet=sheet.title,
                    ))
    finally:
        wb.close()

    return fields


def _is_numeric(text: str) -> bool:
    """True si el texto es puramente numérico (ignora separadores)."""
    clean = re.sub(r"[\s.,\-/]", "", text)
    return bool(clean) and clean.isdigit()


def _is_likely_form_label(text: str) -> bool:
    """Heurística: ¿es este texto probablemente una etiqueta de formulario?"""
    # Termina en dos puntos → definitivamente una etiqueta
    if text.rstrip().endswith(":"):
        return True
    # Texto multi-palabra de longitud razonable → probablemente etiqueta
    if " " in text and 4 <= len(text) <= 120:
        return True
    # Palabras clave cortas conocidas como etiquetas
    normalized = normalize_text(text)
    if normalized in _KNOWN_LABEL_TOKENS:
        return True
    # Contiene algún token clave
    for token in _KNOWN_LABEL_TOKENS:
        if token in normalized:
            return True
    return False


def _has_adjacent_empty_excel(sheet: object, row: int, col: int) -> bool:
    """True si hay al menos una celda vacía adyacente (derecha o abajo)."""
    try:
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        return False

    # Escanear hacia la derecha
    for dc in range(1, _INFER_RIGHT_MAX + 1):
        try:
            cell = sheet.cell(row=row, column=col + dc)  # type: ignore[union-attr]
        except Exception:
            break
        if isinstance(cell, MergedCell):
            continue
        val = cell.value
        if val is None or str(val).strip() == "":
            return True
        # Celda no vacía: detenemos el escaneo hacia la derecha
        break

    # Escanear hacia abajo
    for dr in range(1, _INFER_DOWN_MAX + 1):
        try:
            cell = sheet.cell(row=row + dr, column=col)  # type: ignore[union-attr]
        except Exception:
            break
        if isinstance(cell, MergedCell):
            continue
        val = cell.value
        if val is None or str(val).strip() == "":
            return True
        break

    return False


# ---------------------------------------------------------------------------
# Escáner PDF (campos AcroForm + texto estático)
# ---------------------------------------------------------------------------

# Patrón: "Etiqueta:" al final de línea, o seguida de espacios/guiones bajos
_PDF_LABEL_COLON_RE = re.compile(
    r'^(.{2,100}?)\s*:\s*(?:[_\s]{0,60})?$'
)
# Patrón: "Etiqueta: ______..." (guiones bajos explícitos después del colon)
_PDF_LABEL_UNDERLINE_RE = re.compile(
    r'^(.{2,100}?)\s*:\s*_{3,}'
)


def _scan_pdf(path: Path) -> list[DetectedField]:
    try:
        from pypdf import PdfReader
    except ImportError:
        return []

    try:
        reader = PdfReader(str(path))
        raw_fields = reader.get_fields() or {}
    except Exception:
        return []

    result: list[DetectedField] = []
    seen: set[str] = set()

    for qualified, field_obj in raw_fields.items():
        # Usar nombre local (/T) preferentemente; caer en último segmento del nombre calificado
        local = str(field_obj.get("/T", "")).strip()
        if not local:
            local = qualified.split(".")[-1].strip()
        if not local or local in seen:
            continue
        seen.add(local)
        result.append(DetectedField(
            field_key=label_to_key(local),
            label=local,
            sheet=None,
        ))

    # Si el PDF no tiene campos AcroForm, intentar detección por texto
    if not result:
        result = _scan_pdf_text(reader)

    return result


def _scan_pdf_text(reader: object) -> list[DetectedField]:
    """Detecta campos en PDFs estáticos analizando el texto extraído.

    Busca patrones como:
    - "Etiqueta:"  (al final de línea o seguido de espacios/guiones)
    - "Etiqueta: _______"
    - "Etiqueta1:_____ Etiqueta2:_____"  (múltiples campos por línea)
    """
    result: list[DetectedField] = []
    seen_norm: set[str] = set()

    for page in reader.pages:  # type: ignore[attr-defined]
        try:
            text = page.extract_text() or ""
        except Exception:
            continue

        lines = text.splitlines()
        for line in lines:
            stripped = line.strip()
            if not stripped:
                continue

            # Dividir línea en segmentos (maneja múltiples campos por línea)
            segments = _split_pdf_line_into_segments(stripped)
            for segment in segments:
                label = _extract_pdf_label_from_line(segment)
                if label is None:
                    continue
                if len(label) < _MIN_LABEL_LEN or _is_numeric(label):
                    continue
                if not _is_likely_form_label(label):
                    continue
                norm = normalize_text(label)
                if not norm or norm in seen_norm:
                    continue
                seen_norm.add(norm)
                result.append(DetectedField(
                    field_key=label_to_key(label),
                    label=label,
                    sheet=None,
                ))

    return result


# Separa segmentos cuando hay 3+ guiones bajos seguidos de espacio y más texto
_SEGMENT_SPLIT_RE = re.compile(r'_{3,}\s+(?=\S)')


def _split_pdf_line_into_segments(line: str) -> list[str]:
    """Divide una línea en segmentos cuando contiene múltiples campos separados por guiones.

    "NIT:_______ DV:____ Ciudad:__________"  →  ["NIT:", "DV:", "Ciudad:"]
    """
    parts = _SEGMENT_SPLIT_RE.split(line)
    if len(parts) <= 1:
        return [line]
    # Cada parte excepto la última puede terminar con texto (label)
    # La última parte ya tiene la línea completa del último segmento
    return [p.strip() for p in parts if p.strip()]


def _extract_pdf_label_from_line(line: str) -> str | None:
    """Extrae la etiqueta de una línea de texto con patrón de formulario.

    Ejemplos que coinciden:
      "Nombre o Razón Social:"         → "Nombre o Razón Social"
      "NIT/CC/CE: _______________"     → "NIT/CC/CE"
      "Ciudad:                  "      → "Ciudad"
    """
    # Patrón 1: guiones bajos explícitos después del colon
    m = _PDF_LABEL_UNDERLINE_RE.match(line)
    if m:
        label = m.group(1).strip()
        if label:
            return label

    # Patrón 2: colon al final o seguido de espacios
    m = _PDF_LABEL_COLON_RE.match(line)
    if m:
        label = m.group(1).strip()
        # Descartar si la parte después del colon era texto real (no espacios/guiones)
        after_colon = line[line.index(":") + 1:].strip().replace("_", "").strip()
        if after_colon:
            return None
        if label:
            return label

    return None


# ---------------------------------------------------------------------------
# Escáner Word (tablas)
# ---------------------------------------------------------------------------

def _scan_word(path: Path) -> list[DetectedField]:
    try:
        from docx import Document as DocxDocument
    except ImportError:
        return []

    try:
        doc = DocxDocument(str(path))
    except Exception:
        return []

    fields: list[DetectedField] = []
    seen_norm: set[str] = set()

    for table in doc.tables:
        for row in table.rows:
            cells = row.cells
            for i, cell in enumerate(cells[:-1]):
                text = cell.text.strip()
                if len(text) < _MIN_LABEL_LEN:
                    continue
                if _is_numeric(text):
                    continue
                if not _is_likely_form_label(text):
                    continue
                # La celda siguiente debe estar vacía (campo a rellenar)
                next_text = cells[i + 1].text.strip()
                if next_text:
                    continue
                norm = normalize_text(text)
                if norm in seen_norm:
                    continue
                seen_norm.add(norm)
                fields.append(DetectedField(
                    field_key=label_to_key(text),
                    label=text,
                    sheet=None,
                ))

    return fields
