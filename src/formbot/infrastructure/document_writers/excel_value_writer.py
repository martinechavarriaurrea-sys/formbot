from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from formbot.domain.exceptions import (
    DocumentSaveError,
    PositionOutOfBoundsError,
    ValidationException,
)
from formbot.domain.models import CellPosition

LOGGER = logging.getLogger(__name__)

# Patrón para detectar referencias de rango de celdas en fórmulas de dropdown.
# Ejemplos válidos: $A$1:$A$10, Listas!$A$1:$A$20, 'Mi Hoja'!A1:A10
_RANGE_REF_RE = re.compile(
    r"^(?:'?(?P<sheet>[^'!]+)'?!)?(?P<range>[A-Z$]+[0-9$]+:[A-Z$]+[0-9$]+)$",
    re.IGNORECASE,
)


class ExcelValueWriter:
    def write_value(self, workbook: Workbook, position: CellPosition, value: Any) -> None:
        sheet = self._get_sheet(workbook, position.sheet_name)
        if position.row < 1 or position.column < 1:
            raise PositionOutOfBoundsError(
                f"Posicion invalida fila={position.row}, columna={position.column}"
            )

        target_cell = sheet.cell(row=position.row, column=position.column)
        if isinstance(target_cell, MergedCell):
            merged_range = self._find_merged_range(sheet, position)
            raise PositionOutOfBoundsError(
                "No se puede escribir en una celda combinada no editable: "
                f"{position.sheet_name}!{merged_range}"
            )

        coordinate = f"{get_column_letter(position.column)}{position.row}"
        existing_value = target_cell.value
        if isinstance(existing_value, str) and existing_value.startswith("="):
            LOGGER.warning(
                "La celda %s!%s contiene la fórmula '%s' que será reemplazada por '%s'. "
                "El cálculo automático se perderá.",
                position.sheet_name,
                coordinate,
                existing_value,
                value,
            )

        self._validate_dropdown_value(workbook, sheet, position, value)
        target_cell.value = value

    def save(self, workbook: Workbook, output_path: Path) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            workbook.save(output_path)
        except OSError as exc:
            raise DocumentSaveError(
                f"No fue posible guardar el documento de salida: {output_path}"
            ) from exc

    @staticmethod
    def _get_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
        try:
            return workbook[sheet_name]
        except KeyError as exc:
            raise PositionOutOfBoundsError(f"La hoja '{sheet_name}' no existe") from exc

    @staticmethod
    def _find_merged_range(sheet: Worksheet, position: CellPosition) -> str:
        for merged_range in sheet.merged_cells.ranges:
            if (position.row, position.column) in merged_range.cells:
                return str(merged_range)
        return f"R{position.row}C{position.column}"

    @staticmethod
    def _validate_dropdown_value(
        workbook: Workbook,
        sheet: Worksheet,
        position: CellPosition,
        value: Any,
    ) -> None:
        if value is None:
            return

        data_validations = getattr(sheet.data_validations, "dataValidation", [])
        if not data_validations:
            return

        coordinate = f"{get_column_letter(position.column)}{position.row}"
        for validation in data_validations:
            if getattr(validation, "type", None) != "list":
                continue
            sqref = getattr(validation, "sqref", None)
            if sqref is None or coordinate not in sqref:
                continue

            formula = str(getattr(validation, "formula1", "") or "").strip()

            # Intentar opciones inline primero ("si,no")
            options = ExcelValueWriter._extract_inline_options(formula)

            # Si no son inline, intentar resolver el rango de celdas del workbook
            if options is None:
                options = ExcelValueWriter._extract_range_options(
                    workbook, formula, sheet.title
                )

            if options is None:
                LOGGER.warning(
                    "Dropdown en %s!%s: fórmula '%s' no reconocida como inline ni como rango. "
                    "No se valida el valor '%s'.",
                    position.sheet_name,
                    coordinate,
                    formula,
                    value,
                )
                return

            normalized_value = str(value).strip()
            if normalized_value not in options:
                raise ValidationException(
                    f"Valor '{value}' no permitido para dropdown en "
                    f"{position.sheet_name}!{coordinate}. "
                    f"Opciones validas: {', '.join(options)}"
                )
            return

    @staticmethod
    def _extract_inline_options(formula: str) -> list[str] | None:
        """Extrae opciones de una fórmula inline como '"si,no"' → ['si', 'no']."""
        formula = formula.strip()
        if not (formula.startswith('"') and formula.endswith('"')):
            return None
        options = [item.strip() for item in formula.strip('"').split(",")]
        cleaned = [item for item in options if item]
        return cleaned or None

    @staticmethod
    def _extract_range_options(
        workbook: Workbook,
        formula: str,
        current_sheet_name: str,
    ) -> list[str] | None:
        """Lee las celdas de un rango referenciado en el dropdown y retorna sus valores.

        Soporta referencias como:
          - $A$1:$A$10  (misma hoja)
          - Listas!$A$1:$A$20  (hoja distinta)
          - 'Mi Hoja'!A1:A10  (hoja con espacios)
        """
        # Eliminar el prefijo = si viene como fórmula
        cleaned = formula.lstrip("=").strip()
        match = _RANGE_REF_RE.match(cleaned)
        if not match:
            return None

        sheet_part = match.group("sheet")
        range_part = match.group("range").replace("$", "")
        target_sheet_name = sheet_part.strip("'\"") if sheet_part else current_sheet_name

        if target_sheet_name not in workbook.sheetnames:
            LOGGER.warning(
                "Hoja '%s' referenciada en dropdown no existe en el workbook.",
                target_sheet_name,
            )
            return None

        try:
            target_sheet = workbook[target_sheet_name]
            options: list[str] = []
            for row in target_sheet[range_part]:
                cells = row if hasattr(row, "__iter__") else (row,)
                for cell in cells:
                    if cell.value is not None:
                        val = str(cell.value).strip()
                        if val:
                            options.append(val)
            return options if options else None
        except Exception as exc:
            LOGGER.warning(
                "No se pudo leer el rango '%s' de la hoja '%s' para validar dropdown: %s",
                range_part,
                target_sheet_name,
                exc,
            )
            return None
