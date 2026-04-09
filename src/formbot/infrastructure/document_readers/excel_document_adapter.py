from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from formbot.domain.exceptions import DocumentProcessingError, PositionOutOfBoundsError
from formbot.domain.models import CellPosition
from formbot.domain.ports.document_adapter import DocumentAdapter
from formbot.infrastructure.document_readers.excel_structure_reader import ExcelStructureReader
from formbot.infrastructure.document_writers.excel_value_writer import ExcelValueWriter


class ExcelDocumentAdapter(DocumentAdapter):
    def __init__(self, template_path: Path) -> None:
        self._template_path = template_path
        self._workbook = self._load_template(template_path)
        self._structure_reader = ExcelStructureReader()
        self._value_writer = ExcelValueWriter()

    def find_label(self, text: str, sheet_name: str | None = None) -> CellPosition:
        if sheet_name and sheet_name not in self._workbook.sheetnames:
            raise PositionOutOfBoundsError(f"La hoja '{sheet_name}' no existe")
        return self._structure_reader.find_label(
            workbook=self._workbook,
            text=text,
            sheet_name=sheet_name,
        )

    def write_value(self, position: CellPosition, value: Any) -> None:
        self._value_writer.write_value(self._workbook, position, value)

    def save(self, output_path: Path) -> None:
        self._value_writer.save(self._workbook, output_path)

    def close(self) -> None:
        self._workbook.close()

    def find_adjacent_empty(self, position: CellPosition) -> CellPosition | None:
        return self._structure_reader.find_adjacent_empty(self._workbook, position)

    @staticmethod
    def _load_template(template_path: Path) -> Workbook:
        if not template_path.exists():
            raise DocumentProcessingError(
                f"No existe el template de Excel: {template_path}"
            )
        if template_path.suffix.lower() == ".xls":
            raise DocumentProcessingError(
                f"El formato .xls (Excel 97-2003) no está soportado. "
                f"Convierta el archivo a .xlsx o .xlsm antes de usarlo: {template_path}"
            )
        try:
            return load_workbook(template_path)
        except Exception as exc:  # pragma: no cover - depende del estado real del archivo.
            raise DocumentProcessingError(
                f"No fue posible cargar el template de Excel: {template_path}"
            ) from exc

