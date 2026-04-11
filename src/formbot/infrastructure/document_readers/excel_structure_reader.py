from __future__ import annotations

import datetime
from dataclasses import dataclass

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from formbot.domain.exceptions import LabelNotFoundError, MappingRuleError
from formbot.domain.models import CellPosition
from formbot.shared.utils import normalize_text


@dataclass(frozen=True)
class LabelCandidate:
    position: CellPosition
    match_type: str
    text: str


class ExcelStructureReader:
    def find_label_candidates(
        self,
        workbook: Workbook,
        text: str,
        sheet_name: str | None = None,
    ) -> list[LabelCandidate]:
        normalized_target = normalize_text(text)
        if not normalized_target:
            raise LabelNotFoundError("No se puede buscar una etiqueta vacia")

        sheets = [workbook[sheet_name]] if sheet_name else list(workbook.worksheets)
        exact_matches: list[LabelCandidate] = []
        partial_matches: list[LabelCandidate] = []

        for sheet in sheets:
            self._collect_matches(
                sheet=sheet,
                normalized_target=normalized_target,
                exact_matches=exact_matches,
                partial_matches=partial_matches,
            )

        return [*exact_matches, *partial_matches]

    def find_label(
        self,
        workbook: Workbook,
        text: str,
        sheet_name: str | None = None,
    ) -> CellPosition:
        candidates = self.find_label_candidates(
            workbook=workbook,
            text=text,
            sheet_name=sheet_name,
        )
        exact_matches = [c for c in candidates if c.match_type == "exact"]
        partial_matches = [c for c in candidates if c.match_type == "partial"]

        if len(exact_matches) == 1:
            return exact_matches[0].position
        if len(exact_matches) > 1:
            raise MappingRuleError(
                f"Etiqueta ambigua '{text}'. Se encontraron {len(exact_matches)} coincidencias exactas."
            )
        if len(partial_matches) == 1:
            return partial_matches[0].position
        if len(partial_matches) > 1:
            raise MappingRuleError(
                f"Etiqueta ambigua '{text}'. Se encontraron {len(partial_matches)} coincidencias parciales."
            )
        raise LabelNotFoundError(f"No se encontro etiqueta '{text}' en el documento")

    @staticmethod
    def _collect_matches(
        sheet: Worksheet,
        normalized_target: str,
        exact_matches: list[LabelCandidate],
        partial_matches: list[LabelCandidate],
    ) -> None:
        for row in sheet.iter_rows():
            for cell in row:
                raw = cell.value
                if isinstance(raw, str):
                    cell_text = raw
                elif isinstance(raw, datetime.datetime):
                    # Etiquetas de fecha como "2024-01-15" son poco comunes pero posibles
                    cell_text = raw.strftime("%Y-%m-%d")
                else:
                    continue
                normalized_value = normalize_text(cell_text)
                if not normalized_value:
                    continue

                position = CellPosition(
                    sheet_name=sheet.title,
                    row=cell.row,
                    column=cell.column,
                )
                if normalized_value == normalized_target:
                    exact_matches.append(
                        LabelCandidate(
                            position=position,
                            match_type="exact",
                            text=cell_text,
                        )
                    )
                    continue

                if normalized_target in normalized_value:
                    partial_matches.append(
                        LabelCandidate(
                            position=position,
                            match_type="partial",
                            text=cell_text,
                        )
                    )

    def find_adjacent_empty(
        self, workbook: Workbook, position: CellPosition
    ) -> CellPosition | None:
        """Primera celda vacía adyacente a *position* (derecha hasta 8, luego abajo hasta 3).

        La búsqueda hacia la derecha continúa más allá de celdas ocupadas para
        alinearse con la lógica de PrecisionFillUseCase._infer_target, que evalúa
        todas las posiciones y elige la de mayor score (score=1.0 para celdas vacías).
        """
        try:
            from openpyxl.cell.cell import MergedCell
            sheet = workbook[position.sheet_name]
        except (KeyError, ImportError):
            return None

        # Escanear hacia la derecha (sin parar en la primera celda ocupada)
        for dc in range(1, 9):
            try:
                cell = sheet.cell(row=position.row, column=position.column + dc)
            except Exception:
                break
            if isinstance(cell, MergedCell):
                continue
            if cell.value is None or str(cell.value).strip() == "":
                return CellPosition(position.sheet_name, position.row, position.column + dc)

        # Escanear hacia abajo
        for dr in range(1, 4):
            try:
                cell = sheet.cell(row=position.row + dr, column=position.column)
            except Exception:
                break
            if isinstance(cell, MergedCell):
                continue
            if cell.value is None or str(cell.value).strip() == "":
                return CellPosition(position.sheet_name, position.row + dr, position.column)
            break

        return None
