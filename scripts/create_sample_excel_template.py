from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def main() -> int:
    project_root = Path(__file__).resolve().parents[1]
    template_dir = project_root / "examples" / "templates"
    template_dir.mkdir(parents=True, exist_ok=True)
    template_path = template_dir / "form_template.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Formulario"

    sheet["A1"] = "Nombre del representante legal"
    sheet["D1"] = ""
    sheet["A2"] = "NIT"
    sheet["D2"] = ""
    sheet["A3"] = "Razon social"
    sheet["D3"] = ""
    sheet["A4"] = "Correo electronico"
    sheet["D4"] = ""

    sheet.merge_cells("A6:C6")
    sheet["A6"] = "Bloque de observaciones"
    sheet["D6"] = ""

    title_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    for row in range(1, 7):
        label_cell = sheet[f"A{row}"]
        label_cell.font = Font(bold=True)
        label_cell.fill = title_fill
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet[f"D{row}"].alignment = Alignment(horizontal="left", vertical="center")

    sheet.column_dimensions["A"].width = 36
    sheet.column_dimensions["B"].width = 8
    sheet.column_dimensions["C"].width = 8
    sheet.column_dimensions["D"].width = 36

    workbook.save(template_path)
    workbook.close()
    print(f"Template de ejemplo generado en: {template_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

