"""Verifica que cada offset del YAML apunte a una celda válida en el Excel.

Uso:
    python scripts/verificar_offsets.py <excel_path> [yaml_path]

Si no se indica yaml_path usa config/mappings/cu_for_001_contrapartes.yaml.

Salida:
    - OK-VACÍA   : offset apunta a celda vacía → lista para escribir
    - OK-PLACEHOLDER : apunta a celda con guión/X/placeholder → aceptable
    - OCUPADA    : apunta a celda con valor real → puede sobrescribir datos
    - LABEL?     : apunta a una celda que parece otra etiqueta → offset malo
    - FUERA-RANGO: la posición calculada está fuera de la hoja
    - LABEL-NO-ENCONTRADO: el label del YAML no existe en el Excel
    - AMBIGUO    : el label aparece más de una vez → necesita hint_row o alias más específico
    - SKIP       : campo sin strategy=offset (mark / infer) → no se verifica offset
"""
from __future__ import annotations

import sys
from pathlib import Path

# Asegurar que src/ esté en el path para poder importar formbot.shared.utils
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

import re
from dataclasses import dataclass

import yaml

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
except ImportError:
    print("ERROR: openpyxl no instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)

try:
    from formbot.shared.utils import normalize_text
except ImportError:
    def normalize_text(text: str) -> str:  # type: ignore[misc]
        import unicodedata
        nfkd = unicodedata.normalize("NFKD", text.lower())
        return "".join(c for c in nfkd if not unicodedata.combining(c)).strip()


# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

_PLACEHOLDER_CHARS: frozenset[str] = frozenset({"_", "-", ".", "x", " "})
_LABEL_MARKERS: tuple[str, ...] = (
    "nombre", "razon social", "direccion", "ciudad", "telefono",
    "correo", "representante", "firma", "fecha", "nit", "numero",
    "identificacion", "documento", "tipo", "cuenta", "cargo",
    "contacto", "banco", "entidad", "titular",
)

_STATUS_WIDTH = 22


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

@dataclass
class VerifyResult:
    field_name: str
    status: str
    detail: str
    label_found: str | None = None
    target_coord: str | None = None
    target_value: str | None = None
    occurrences: int = 0


def _is_placeholder(value: str) -> bool:
    if not value:
        return True
    stripped = value.strip()
    return bool(stripped) and set(stripped.lower()) <= _PLACEHOLDER_CHARS


def _looks_like_label(value: str) -> bool:
    if ":" in value:
        return True
    norm = normalize_text(value)
    return any(m in norm for m in _LABEL_MARKERS)


def _cell_status(cell_value) -> tuple[str, str]:
    """Devuelve (status, display_value) para una celda objetivo."""
    if cell_value is None:
        return "OK-VACÍA", ""
    text = str(cell_value).strip()
    if not text:
        return "OK-VACÍA", ""
    if _is_placeholder(text):
        return "OK-PLACEHOLDER", text[:40]
    if _looks_like_label(text):
        return "LABEL?", text[:40]
    return "OCUPADA", text[:40]


def _find_label_cells(
    ws,
    label_text: str,
    aliases: list[str],
) -> list[tuple[int, int, str]]:
    """Encuentra todas las celdas que coincidan con label o sus aliases."""
    search_terms = [label_text] + aliases
    norm_terms = [normalize_text(t) for t in search_terms]
    hits: list[tuple[int, int, str]] = []

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is None:
                continue
            cell_norm = normalize_text(str(cell.value))
            for norm in norm_terms:
                if norm and (cell_norm == norm or cell_norm.startswith(norm)):
                    hits.append((cell.row, cell.column, str(cell.value).strip()))
                    break

    return hits


# ---------------------------------------------------------------------------
# Verificador principal
# ---------------------------------------------------------------------------

def verify(excel_path: Path, yaml_path: Path) -> list[VerifyResult]:
    rules: dict = yaml.safe_load(yaml_path.read_text(encoding="utf-8"))

    wb = load_workbook(excel_path, data_only=True)
    results: list[VerifyResult] = []

    for field_name, rule in rules.items():
        strategy = rule.get("target_strategy", "offset")
        sheet_name: str = rule.get("sheet", "")
        label_text: str = rule.get("label", "")
        aliases: list[str] = rule.get("aliases", [])

        # Campos de marcación sin offset fijo → solo informar
        if strategy in {"infer", "mark"} and rule.get("target_strategy") != "offset":
            results.append(VerifyResult(
                field_name=field_name,
                status="SKIP",
                detail=f"strategy={strategy} — sin offset fijo, no aplica verificación",
            ))
            continue

        if strategy == "offset_or_infer" and not rule.get("offset"):
            results.append(VerifyResult(
                field_name=field_name,
                status="SKIP",
                detail="offset_or_infer sin offset configurado — usa inferencia",
            ))
            continue

        # Verificar que la hoja exista
        if sheet_name not in wb.sheetnames:
            # Intentar búsqueda insensible a mayúsculas/tildes
            norm_sheet = normalize_text(sheet_name)
            matched = next(
                (s for s in wb.sheetnames if normalize_text(s) == norm_sheet), None
            )
            if matched is None:
                results.append(VerifyResult(
                    field_name=field_name,
                    status="HOJA-NO-ENCONTRADA",
                    detail=f"Hoja '{sheet_name}' no existe. Disponibles: {wb.sheetnames}",
                ))
                continue
            sheet_name = matched

        ws = wb[sheet_name]

        # Leer offset
        offset_cfg = rule.get("offset", {})
        row_off: int = offset_cfg.get("row", 0)
        col_off: int = offset_cfg.get("col", 1)

        # Buscar etiqueta en la hoja
        hits = _find_label_cells(ws, label_text, aliases)

        if not hits:
            results.append(VerifyResult(
                field_name=field_name,
                status="LABEL-NO-ENCONTRADO",
                detail=f"'{label_text}' no encontrado en hoja '{sheet_name}'",
                occurrences=0,
            ))
            continue

        if len(hits) > 1:
            has_hint = rule.get("hint_row") is not None
            if not has_hint:
                coords = [f"R{r}C{c}" for r, c, _ in hits]
                results.append(VerifyResult(
                    field_name=field_name,
                    status="AMBIGUO",
                    detail=f"Label encontrado {len(hits)}x: {coords} — necesita hint_row o alias más específico",
                    label_found=hits[0][2],
                    occurrences=len(hits),
                ))
                continue
            # Con hint_row: usar la instancia más cercana
            hint = rule["hint_row"]
            hits.sort(key=lambda h: abs(h[0] - hint))

        label_row, label_col, label_found = hits[0]
        target_row = label_row + row_off
        target_col = label_col + col_off

        # Verificar que la posición target esté dentro de la hoja
        if target_row < 1 or target_col < 1:
            results.append(VerifyResult(
                field_name=field_name,
                status="FUERA-RANGO",
                detail=f"Label en R{label_row}C{label_col} + offset({row_off},{col_off}) → R{target_row}C{target_col} (inválido)",
                label_found=label_found,
                occurrences=len(hits),
            ))
            continue

        # Verificar la celda target
        try:
            target_cell = ws.cell(row=target_row, column=target_col)
            if isinstance(target_cell, MergedCell):
                status = "MERGED"
                display = "(celda fusionada)"
            else:
                status, display = _cell_status(target_cell.value)
        except Exception as exc:
            status = "ERROR"
            display = str(exc)

        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(target_col)
        coord = f"{col_letter}{target_row}"

        results.append(VerifyResult(
            field_name=field_name,
            status=status,
            detail=f"Label R{label_row}C{label_col} + ({row_off},{col_off}) → {coord}",
            label_found=label_found,
            target_coord=coord,
            target_value=display or None,
            occurrences=len(hits),
        ))

    wb.close()
    return results


# ---------------------------------------------------------------------------
# Impresión y resumen
# ---------------------------------------------------------------------------

_STATUS_COLORS = {
    "OK-VACÍA": "\033[32m",        # verde
    "OK-PLACEHOLDER": "\033[32m",  # verde
    "SKIP": "\033[90m",            # gris
    "OCUPADA": "\033[33m",         # amarillo
    "LABEL?": "\033[33m",          # amarillo
    "AMBIGUO": "\033[33m",         # amarillo
    "MERGED": "\033[33m",          # amarillo
    "LABEL-NO-ENCONTRADO": "\033[31m",   # rojo
    "HOJA-NO-ENCONTRADA": "\033[31m",    # rojo
    "FUERA-RANGO": "\033[31m",           # rojo
    "ERROR": "\033[31m",                 # rojo
}
_RESET = "\033[0m"


def _supports_color() -> bool:
    import os
    return hasattr(sys.stdout, "isatty") and sys.stdout.isatty() and os.name != "nt"


def print_results(results: list[VerifyResult]) -> None:
    use_color = _supports_color()
    sep = "-" * 120
    print()
    print(f"{'CAMPO':<42} {'STATUS':<22} DETALLE")
    print(sep)

    counts: dict[str, int] = {}
    for r in results:
        counts[r.status] = counts.get(r.status, 0) + 1
        color = _STATUS_COLORS.get(r.status, "") if use_color else ""
        reset = _RESET if use_color else ""
        value_hint = f" -> '{r.target_value}'" if r.target_value else ""
        print(
            f"{r.field_name:<42} "
            f"{color}{r.status:<22}{reset} "
            f"{r.detail}{value_hint}"
        )

    print(sep)
    print("\nResumen:")
    for status, count in sorted(counts.items()):
        color = _STATUS_COLORS.get(status, "") if use_color else ""
        reset = _RESET if use_color else ""
        print(f"  {color}{status:<22}{reset} {count}")

    errors = sum(
        v for k, v in counts.items()
        if k in {"LABEL-NO-ENCONTRADO", "HOJA-NO-ENCONTRADA", "FUERA-RANGO", "AMBIGUO", "ERROR"}
    )
    warnings = sum(
        v for k, v in counts.items()
        if k in {"OCUPADA", "LABEL?", "MERGED"}
    )
    print()
    if errors == 0 and warnings == 0:
        print("OK Sin errores ni advertencias detectados.")
    else:
        if errors:
            print(f"ERROR {errors} error(s) requieren correccion en el YAML.")
        if warnings:
            print(f"WARN  {warnings} advertencia(s) — revisar manualmente.")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    excel_path = Path(sys.argv[1])
    yaml_path = (
        Path(sys.argv[2])
        if len(sys.argv) >= 3
        else Path(__file__).resolve().parents[1]
        / "config/mappings/cu_for_001_contrapartes.yaml"
    )

    if not excel_path.exists():
        print(f"ERROR: No existe el archivo Excel: {excel_path}")
        sys.exit(1)
    if not yaml_path.exists():
        print(f"ERROR: No existe el YAML: {yaml_path}")
        sys.exit(1)

    print(f"Excel : {excel_path}")
    print(f"YAML  : {yaml_path}")

    results = verify(excel_path, yaml_path)
    print_results(results)

    errors = sum(
        1 for r in results
        if r.status in {"LABEL-NO-ENCONTRADO", "HOJA-NO-ENCONTRADA", "FUERA-RANGO", "AMBIGUO", "ERROR"}
    )
    sys.exit(1 if errors > 0 else 0)


if __name__ == "__main__":
    main()
